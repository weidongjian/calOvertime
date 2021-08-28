# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

import openpyxl
import datetime
import chinese_calendar

# 一天工作多少秒
TOTAL_WORK_SECOND = 7.5 * 60 * 60
LAUNCH_BREAK_SECOND = 1.5 * 60 * 60

COLUMN_DATE = 0
COLUMN_WEEK = 1
COLUMN_NAME = 2
COLUMN_ID = 3
COLUMN_TIME = 8


class workItem:
    def __init__(self):
        self.startTime = ""
        self.endTime = ""

    def reset(self):
        self.startTime = ""
        self.idValue = ""


class UserItem:
    def __init__(self, name, idValue):
        self.name = name
        self.idValue = idValue
        self.dict = {}  # 打卡信息，key是日期，value是打卡list


# 总的加班信息,key是员工账号， value是UserItem
workMap = {}


def isWeekend(dataStr):
    date = datetime.datetime.strptime(dataStr, "%Y/%m/%d")
    return chinese_calendar.is_holiday(date)


def calOverTimeReal(startTime, endTime):
    try:
        timeOne = datetime.datetime.strptime(startTime, '%H:%M')
        timeTwo = datetime.datetime.strptime(endTime, '%H:%M')
    except ValueError:
        return 0

    timeDiffSec = (timeTwo - timeOne).total_seconds()
    if timeOne.hour < 12:
        # 减去午休的时间
        timeDiffSec = timeDiffSec - LAUNCH_BREAK_SECOND

    return timeDiffSec


def getUserItemFromMap(idValue, nameValue):
    if idValue in workMap.keys():
        return workMap[idValue]
    else:
        valueItem = UserItem(nameValue, idValue)
        workMap[idValue] = valueItem
        return valueItem


def getWorkListFromUser(user_item, dateValue):
    if dateValue in user_item.dict.keys():
        return user_item.dict[dateValue]
    else:
        workList = []
        user_item.dict[dateValue] = workList
        return workList


def updateWorkData(user_item, dateValue, timeValue):
    workList = getWorkListFromUser(user_item, dateValue)
    workList.append(timeValue)


def calTotalWorkTime(workDic):
    totalSecond = 0
    keyList = []
    for key in workDic.keys():
        keyList.append(key)
        workList = workDic[key]
        if len(workList) >= 2:
            startTime = workList[0]
            endTime = workList[-1]
            time_real = calOverTimeReal(startTime, endTime)
            totalSecond = totalSecond + time_real

    totalMin = totalSecond // 60
    totalHour = totalMin // 60
    remainMin = totalMin % 60
    workResult = str(int(totalHour)) + ":" + str(int(remainMin))
    mat = "{:6}\t{:10}"
    mat_format = mat.format("总加班时长", workResult)
    print(mat_format + " 加班日期 " + str(keyList))


def calculateResult():
    for value in workMap.values():
        mat = "{:10}\t{:12}"
        print(mat.format(value.name, value.idValue), end="")
        workDic = value.dict
        calTotalWorkTime(workDic)


def main():
    sourceFile = input("请输入考勤文件地址：")
    # sourceFile = "/Users/weigan/Downloads/7月考勤全员.xlsx"
    workbook = openpyxl.load_workbook(sourceFile.strip(), read_only=True)
    sheet = workbook.worksheets[1]
    sheet.iter_rows(min_row=5)
    for row in sheet.iter_rows(min_row=5, values_only=True):
        dateValue = row[COLUMN_DATE]
        weekValue = row[COLUMN_WEEK]
        nameValue = row[COLUMN_NAME]
        idValue = row[COLUMN_ID]
        timeValue = row[COLUMN_TIME]
        if not isWeekend(dateValue):
            continue
        user_item = getUserItemFromMap(idValue, nameValue)
        updateWorkData(user_item, dateValue, timeValue)

    calculateResult()


if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
